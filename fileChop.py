import openpyxl as opxl
import queryAI
from openpyxl.utils import column_index_from_string
from tkinter.filedialog import askopenfilename
import os

def main():
    filename = askopenfilename(title="Select an EXCEL workbook",
                               filetypes=[("Excel files", ".xlsx")])  # openpyxl also works on xlsm files, but let's just keep it simple for now
    if not filename: exit()
    filenameSansExt, ext = os.path.splitext(filename)

    wb = opxl.Workbook()
    wb = opxl.load_workbook(filename = filename)
    
    print("Available worksheets:")
    for i in range(len(wb.sheetnames)):
        print(f"\t{i} -- {wb.sheetnames[i]}")
    print()
    
    # ERROR: If user inputs an index outside the range of the list.
    selectedWsIndex = input("Select index number: ")
    # ERROR: If user inputs a non-integer.
    selectedWsIndex = int(selectedWsIndex)
    selWs = wb[wb.sheetnames[selectedWsIndex]]
    print(f"You picked '{selWs.title}'")
    print()

    selectedColumnLetter = input("Select a column (from 'A' to 'XFD'): ")
    selectedColumnLetter = selectedColumnLetter.strip().upper()
    
    colIndex = column_index_from_string(selectedColumnLetter)
    # ERROR: If user inputs something outside the range A to XFD (1 - 16384)
    print(f"Column '{selectedColumnLetter}' corresponds to index {colIndex}.")
    print()

    selectedFirstRow = int(input("Select first row: "))
    selectedLastRow = int(input("Select last row (inclusive): "))
    numItems = selectedLastRow - selectedFirstRow + 1
    print(f"Ranging from rows {selectedFirstRow} to {selectedLastRow} ({numItems} items).")
    print()



    cellsList = []
    # IMPROPER: The max_row attribute of a worksheet does not take the column into account. Cannot get max row by worksheet column.
    for row in selWs.iter_rows(min_row=selectedFirstRow, max_row=selectedLastRow, min_col=colIndex, max_col=colIndex):
        cellsList.append(row[0].value)
        # rowValue = row[0].value
        # if not rowValue: print()
        # else: print(rowValue)
    
    print("List of values in this column:")
    print(cellsList)
    print()


    paragraphList = queryAI.queryAI(str(cellsList)).output_parsed
    # for paragraph in paragraphList.correctedParagraphs:
    #     if paragraph:
    #         print(paragraph)
    #     else:
    #         print("\"\" ")
    
    i = 0
    nextColIndex = colIndex + 1
    for cellRow in range(selectedFirstRow, selectedLastRow+1):
        selWs.cell(row=cellRow, column=nextColIndex, value=paragraphList.correctedParagraphs[i])
        i += 1


    revisedFilename = f"{filenameSansExt}-Revised{ext}"
    wb.save(revisedFilename)
    os.startfile(revisedFilename)
    
    return


    # aiResponse = aiScript.queryAI(str(cellsList))
    # print(aiResponse)
    # aiResponseList = aiResponse.split(",")
    # print(aiResponseList)

    # for i in range(len(aiResponseList)):
    #     aiResponseList[i] = aiResponseList[i].strip()
    #     ws0[f'B{i+1}'] = aiResponseList[i]

    # ws0.column_dimensions['A'].width = 20
    # ws0.column_dimensions['B'].width = 20

    # filenameRevised = filenameSansExt + "-Revised" + ".xlsx"
    # wb.save(filenameRevised)
    # os.startfile(filenameRevised)


if __name__ == "__main__":
    main()