import openpyxl as opxl
# import shutil
import aiScript
from tkinter.filedialog import askopenfilename
import os

filename = askopenfilename(filetypes=[("Excel files", ".xlsx")])
filenameSansExt, ext = os.path.splitext(filename)

wb = opxl.Workbook()

wb = opxl.load_workbook(filename = filename)
ws0 = wb.active

# for row in ws0['A1':'A5']:
#     for cell in row:
#         print(cell.value)

cellsList = []
for row in ws0.iter_rows(min_row=0, max_col=1, max_row=10):
    for cell in row:
        if cell.value:
            cellsList.append(cell.value)

aiResponse = aiScript.queryAI(str(cellsList))
print(aiResponse)
aiResponseList = aiResponse.split(",")
print(aiResponseList)

for i in range(len(aiResponseList)):
    aiResponseList[i] = aiResponseList[i].strip()
    ws0[f'B{i+1}'] = aiResponseList[i]

ws0.column_dimensions['A'].width = 20
ws0.column_dimensions['B'].width = 20

filenameRevised = filenameSansExt + "-Revised" + ".xlsx"
wb.save(filenameRevised)
os.startfile(filenameRevised)
