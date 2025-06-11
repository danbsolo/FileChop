from openpyxl.styles import Font, PatternFill
# from openpyxl.utils import get_column_letter
import os



def exceedCharacterLimitSuggesterFunction(inColIndex, outColIndex, firstRow, lastRow, inputList, outputList, ws):
    i = 0
    for cellRow in range(firstRow, lastRow+1):
        ws.cell(row=cellRow, column=outColIndex, value=outputList.outputList[i]) # col E
        i += 1



def fileNinjaModifierFunction(inColIndex, outColIndex, firstRow, lastRow, inputList, outputList, ws):
    """Slight bug in that a file can be successfully 'renamed' into the same text but with different cases,
    even if it wasn't actually modified."""

    # Always assumes the directory index is column 0
    directoryColIndex = 1 # inColIndex -1
    # oldFilenameColIndex = inColIndex
    # newFilenameColIndex = outColIndex

    # define font stuff
    boldFont = Font(bold=True) # grayish
    successFill = PatternFill(fill_type="solid", fgColor="00FF80")
    failureFill = PatternFill(fill_type="solid", fgColor="FF4444")

    # get the first directory to be used    
    currentDirectory = ""
    i = firstRow
    while not currentDirectory:
        if (potentialDirectory := ws.cell(row=i, column=directoryColIndex).value):
            currentDirectory = potentialDirectory
        i -= 1


    for cellRow in range(firstRow, lastRow+1):
        # if we're on a new directory now, update currentDirectory
        if (potentialDirectory := ws.cell(row=cellRow, column=directoryColIndex).value):
            currentDirectory = potentialDirectory

        oldFilenameCell = ws.cell(row=cellRow, column=inColIndex)
        newFilenameCell = ws.cell(row=cellRow, column=outColIndex)

        # if oldFilenameCell and newFilenameCell are both not none and not identical, make the change
        if (oldFilenameCell.value and newFilenameCell.value) and (oldFilenameCell.value != newFilenameCell.value):
            try:
                os.rename(
                    f"{currentDirectory}\\{oldFilenameCell.value}",
                    f"{currentDirectory}\\{newFilenameCell.value}"
                )

                newFilenameCell.font = boldFont
                newFilenameCell.fill = successFill

            except Exception as e:
                newFilenameCell.font = boldFont
                newFilenameCell.fill = failureFill
        
        # write some sort of indicator the status of the rename in the new excel sheet



def typoFixerFunction(inColIndex, outColIndex, firstRow, lastRow, inputList, outputList, ws):
    # write
    i = 0
    for cellRow in range(firstRow, lastRow+1):
        ws.cell(row=cellRow, column=outColIndex, value=outputList.outputList[i])
        i += 1
